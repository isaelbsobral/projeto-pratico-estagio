from openpyxl import Workbook#importando modulo da biblioteca openpyxl para criação de uma pasta de trabalho
arq_excel = Workbook() #instancia uma pasta de trabalho
planilha4 = arq_excel.active #cria e ativa a primeira planilha da pasta de trabalho

planilha4.title = "Links"

print(arq_excel.sheetnames)#retorna a lista com todas as planilhas
#planilha4 = arq_excel.create_sheet("Links")
planilha4.cell(row=1, column=1, value='link')
planilha4.cell(row=1, column=2, value='atualTime')

dicta= {'https://blog.letscode.com.br': '18:49:00', 'https://www.letscode.com.br/': '18:49:00', 'https://www.facebook.com/letscodebr/': '18:49:00', 'https://openpyxl.readthedocs.io/en/stable/': '18:49:00', 'https://ghost.org/': '18:49:00'}
dict1 = list(dicta.keys())
dict2 = list(dicta.values())
print(len(dicta))
print(dict1)
print(dict2)
for i in range(1,len(dicta)+1):
    planilha4.cell(row=i+1, column=1, value=dict1[i-1])
    planilha4.cell(row=i+1, column=2, value=dict2[i-1])
    print(i)
#print(dicta.keys()[1])
#print(dicta.values()[1])
#for i in dicta:


arq_excel.save("planilha4.xlsx")
