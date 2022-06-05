from openpyxl import Workbook#importando módulo da biblioteca openpyxl para criação de uma pasta de trabalho
import requests
from bs4 import BeautifulSoup
from requests.exceptions import HTTPError
from datetime import datetime #Biblicoteca para manipular data e hora
from openpyxl.styles import Font,Color, colors, Alignment #Importando módulos para estilizar a planilha.
dicionarioLinkseHorarios={}
def getLinks(urlArgumento, qtdNiveis, nomeArquivo):
    try:  #Tenta fazer a conexão com a página WEB
        paginaWeb = requests.get(
            urlArgumento)  # Aqui o método está fazendo a leitra da URL recebida e armazenando em variável.
        paginaWeb.raise_for_status()  # Se a requisição da página foi feita com sucesso ele não entra, mas executa as exceções.

    except HTTPError as erroHTTP:  # Se não houve sucesso na requisição ele executa e exibe o erro.
        print("")
        print(f'Um erro de requisição HTTP ocorreu:\n{erroHTTP}')
        print("Tente novamente!")
        getLinks(insereURL(), qtdNiveis, nomeArquivo)  # Solicita do usuário uma URL válida e usa recursão para tentar conectar novamente.
    except Exception as erro:
        print("")
        print(f'Um erro de requisição aconteceu: \n{erro}')
        print("Tente novamente!")
        getLinks(insereURL(), qtdNiveis, nomeArquivo)
    else:
        print("")
        print('Conexão realizada com sucesso!')
        paginaWebConvertidaBS = BeautifulSoup(paginaWeb.text,
                                              'html.parser')  # Converte o conteúdo da página web em um objeto BeautifulSoup.
        listaComResultadoLinks = paginaWebConvertidaBS.find_all('a')
        for i in listaComResultadoLinks:#Percorre a lista encontrada.
            # Captura e armazena os links das urls válidas dentro do atributo href
            linkHref = i.get("href")
            if linkHref != None and linkHref.startswith('http') == True: #Verifica se a string com link encontrada tem valor vazio e se começa com http.
                #Preenche um dicionário com o link e o horário que ele foi encontrado.
                linkEncontrado = linkHref
                horarioAtual = datetime.now().time().strftime('%H:%M:%S')#Pega o horário atual do computador, converte para o formato H:M:S e armazena dentro da lista após o link ser encontrado.
                dicionarioLinkseHorarios[linkEncontrado] = horarioAtual
   # print(dicionarioLinkseHorarios) imprime no terminal os valores armazenados no dicionário

    minha_pastaTrabalho = Workbook() #Instancia uma pasta de trabalho.
    minha_planilha = minha_pastaTrabalho.active #Cria e ativa a primeira planilha da pasta de trabalho
    minha_planilha.title = "Links" #Nomeando o título da planilha.

    #Nomeando os títulos da coluna.
    minha_planilha.cell(row=1, column=1, value="link")
    minha_planilha.cell(row=1, column=2, value="atualTime")
    '''
    Tentativa de estilizar a planilha via código
    formatCol1 = minha_planilha['A1']
    formatCol2 = minha_planilha['A2']
    formatCol1 = Font(size=14, bold=True)
    formatCol2 = Font(size=14, bold=True)
    formatCol1 = Alignment(vertical='center')
    formatCol2 = Alignment(vertical='center')'''
    #Armazena em lista diferentes os links e os horários do dicionário
    listaChaves = list(dicionarioLinkseHorarios.keys())
    listaValores = list(dicionarioLinkseHorarios.values())
    #Armazena na coluna 'link' os links contidos na listaChaves e na coluna 'atualTime' os horários contidos na listaValores.
    for j in range(1, len(dicionarioLinkseHorarios)+1):
        minha_planilha.cell(row=j+1, column=1, value=listaChaves[j-1])
        minha_planilha.cell(row=j+1, column=2, value=listaValores[j-1])

    minha_pastaTrabalho.save(nomeArquivo) #Salva a planilha com o nome informado pelo usuário.

def insereURL():
    print("Cole abaixo a URL:")
    urlRequisitada = input()
    return urlRequisitada
print("Sistema para salvar links de uma página WEB em uma planilha excel")
nomeArquivo = input("Informe como deseja salvar o nome do arquivo(formato:nomeArquivo.xlsx:")
qtdNiveis = int(input("Digite em quantos níveis deseja fazer a procura:")) #Não consegui implementar
#Chamada da função
getLinks(insereURL(), qtdNiveis, nomeArquivo)

